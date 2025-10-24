#!/usr/bin/env python3
import sys
from pathlib import Path
from typing import List, Tuple

try:
    from openpyxl import load_workbook
except ImportError as e:
    print("openpyxl is not installed. Install with: python3 -m pip install --user openpyxl", file=sys.stderr)
    sys.exit(1)

TEMPLATE_DEFAULT = Path('frontend/public/templates/Social_Garden_SOW_Template.xlsx')

HEADERS_SUMMARY = {"SCOPE", "TOTAL HOURS", "SUBTOTAL (EX. GST)", "GST (10%)", "TOTAL (INC. GST)"}
HEADERS_SCOPE = {"ROLE", "HOURS", "RATE (AUD)", "TOTAL (AUD)"}


def normalize(v: str) -> str:
    return (v or "").strip().upper()


def find_header_anchor(ws, expected: set) -> Tuple[int, int]:
    """Return (row, col) of header's top-left cell if a row contains most expected headers."""
    max_rows = min(ws.max_row, 100)
    max_cols = min(ws.max_column, 20)
    for r in range(1, max_rows + 1):
        row_vals = []
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(normalize(str(v) if v is not None else ""))
        match = sum(1 for h in expected if h in row_vals)
        if match >= max(3, len(expected) - 1):
            # return first column where the first expected header appears
            for c, v in enumerate(row_vals, start=1):
                if v in expected:
                    return r, c
    return -1, -1


def last_data_row(ws, start_row: int, key_col: int) -> int:
    """Heuristic: walk down from start_row+1 until a few consecutive empty rows are found."""
    if start_row < 1:
        return start_row
    r = start_row + 1
    empty_streak = 0
    last = start_row
    max_r = ws.max_row

    while r <= max_r and empty_streak < 5:
        v = ws.cell(row=r, column=key_col).value
        if v is None or str(v).strip() == "":
            empty_streak += 1
        else:
            empty_streak = 0
            last = r
        r += 1
    return last


def print_defined_names(wb):
    try:
        names = list(wb.defined_names.definedName)
    except Exception:
        names = []
    if not names:
        print("Defined Names: (none)")
        return
    print("Defined Names:")
    for dn in names:
        print(f"  - {dn.name}")
        try:
            dests = dn.destinations
            for title, coord in dests:
                print(f"      -> {title}!{coord}")
        except Exception:
            pass


def main(path: str = None):
    tpl = Path(path) if path else TEMPLATE_DEFAULT
    if not tpl.exists():
        print(f"Template not found at {tpl.resolve()}", file=sys.stderr)
        sys.exit(2)

    wb = load_workbook(tpl)
    print(f"Loaded template: {tpl}")
    print("Sheets:", ", ".join(ws.title for ws in wb.worksheets))
    print_defined_names(wb)

    # Summary detection
    summary = wb["SOW_Summary"] if "SOW_Summary" in wb.sheetnames else None
    if summary:
        sr, sc = find_header_anchor(summary, HEADERS_SUMMARY)
        if sr != -1:
            print(f"Summary table header anchor: {summary.title}!{summary.cell(row=sr, column=sc).coordinate}")
            last = last_data_row(summary, sr, sc)
            print(f"Summary table last detected row (heuristic): {last}")
        else:
            print("Summary header not detected (using defaults: header at row 3, A3:E3)")

    # Scope sheets
    for ws in wb.worksheets:
        if not ws.title.lower().startswith("scope"):
            continue
        r, c = find_header_anchor(ws, HEADERS_SCOPE)
        if r != -1:
            print(f"{ws.title} pricing header anchor: {ws.title}!{ws.cell(row=r, column=c).coordinate}")
            last = last_data_row(ws, r, c)
            print(f"{ws.title} pricing last detected row (heuristic): {last}")
        else:
            print(f"{ws.title}: pricing header not detected (using defaults: header at row 5, A5:D5)")

    print("\nSuggested Named Ranges (add these in Excel for precise placement):")
    print("  SUMMARY_TITLE -> SOW_Summary!A1 (or your actual title cell)")
    print("  SUMMARY_TABLE_START -> SOW_Summary!A3")
    print("  For each scope sheet:")
    print("    SCOPE#_TITLE -> Scope#(!)A1")
    print("    SCOPE#_OVERVIEW -> Scope#(!)A3")
    print("    SCOPE#_TABLE_START -> Scope#(!)A5")
    print("    SCOPE#_DELIVERABLES_START -> Scope#(!)A{after-totals}")
    print("    SCOPE#_ASSUMPTIONS_START -> Scope#(!)A{after-deliverables}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else None)
